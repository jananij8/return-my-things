from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_mysqldb import MySQL
from config import Config
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
import os

app = Flask(__name__)
app.config.from_object(Config)

app.config["MYSQL_CURSORCLASS"] = "DictCursor"
mysql = MySQL(app)

app.secret_key = Config.SECRET_KEY

# ==========================
# HOME PAGE
# ==========================

@app.route("/")
def home():
    return render_template("login.html")


# ==========================
# REGISTER PAGE
# ==========================

@app.route("/register")
def register_page():
    return render_template("register.html")
# ==========================
# LOGIN
# ==========================

@app.route("/login", methods=["POST"])
def login():

    role = request.form["role"]
    username = request.form["username"]
    password = request.form["password"]

    cursor = mysql.connection.cursor()

    # ----------------------
    # STUDENT LOGIN
    # ----------------------
    if role == "student":

        cursor.execute(
            "SELECT * FROM students WHERE register_number=%s",
            (username,)
        )

        student = cursor.fetchone()

        if not student:
            flash("You don't have an account. Please register first!", "danger")
            return redirect("/")

        if student["password"] != password:
            flash("Invalid Password!", "danger")
            return redirect("/")

        session["student_id"] = student["student_id"]
        session["student_name"] = student["student_name"]
        return redirect("/student_dashboard")

    # ----------------------
    # STAFF LOGIN
    # ----------------------
    elif role == "staff":

        cursor.execute(
            "SELECT * FROM staff WHERE staff_code=%s",
            (username,)
        )

        staff = cursor.fetchone()

        if staff:

            if staff["password"] == password:

                session["staff_id"] = staff["staff_id"]
                session["staff_name"] = staff["staff_name"]

                return redirect("/staff_dashboard")

        flash("Invalid Staff Login")
        return redirect("/")

    # ----------------------
    # ADMIN LOGIN
    # ----------------------
    elif role == "admin":

        cursor.execute(
            "SELECT * FROM admins WHERE username=%s",
            (username,)
        )

        admin = cursor.fetchone()

        if admin:

            if admin["password"] == password:

                session["admin_id"] = admin["admin_id"]
                session["admin_name"] = admin["full_name"]

                return redirect("/admin_dashboard")

        flash("Invalid Admin Login")
        return redirect("/")

    return redirect("/")
# ==========================
# STUDENT REGISTRATION
# ==========================

@app.route("/register", methods=["POST"])
def register():

    student_name = request.form["student_name"]
    register_number = request.form["register_number"]
    email = request.form["email"]
    phone = request.form["phone"]
    department = request.form["department"]
    year = request.form["year"]
    password = request.form["password"]
    confirm_password = request.form["confirm_password"]

    if password != confirm_password:
        flash("Passwords do not match!","wrong password")
        return redirect("/register")

    cursor = mysql.connection.cursor()

    # Check if register number already exists
    cursor.execute(
        "SELECT * FROM students WHERE register_number=%s",
        (register_number,)
    )

    existing_student = cursor.fetchone()

    if existing_student:
        flash("Register Number already exists!","already exists")
        return redirect("/register")

    # Get department ID
    cursor.execute(
        "SELECT department_id FROM departments WHERE department_name=%s",
        (department,)
    )

    department_data = cursor.fetchone()
    print("Department selected:", department)
    print("Department data:", department_data)

    if not department_data:
        flash("Invalid Department!","invalid department")
        return redirect("/register")

    department_id = department_data["department_id"]

    # Insert Student
    cursor.execute("""
        INSERT INTO students
        (register_number, student_name, department_id,
         year_of_study, email, phone, password)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
    """,
    (
        register_number,
        student_name,
        department_id,
        year,
        email,
        phone,
        password
    ))

    mysql.connection.commit()

    flash("Registration Successful! Please login.", "success")

    return redirect("/")
# ==========================
# STUDENT DASHBOARD
# ==========================

@app.route("/student_dashboard")
def student_dashboard():

    if "student_id" not in session:
        return redirect("/")

    return render_template("student_dashboard.html")


# ==========================
# STAFF DASHBOARD
# ==========================

@app.route("/staff_dashboard")
def staff_dashboard():

    if "staff_id" not in session:
        return redirect("/")

    return render_template("staff_dashboard.html")


# ==========================
# ADMIN DASHBOARD
# ==========================

@app.route("/admin_dashboard")
def admin_dashboard():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM students")
    total_students = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM staff")
    total_staff = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM lost_items")
    total_lost = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM found_items")
    total_found = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM claim_requests")
    total_claims = cursor.fetchone()["total"]

    return render_template(
        "admin_dashboard.html",
        total_students=total_students,
        total_staff=total_staff,
        total_lost=total_lost,
        total_found=total_found,
        total_claims=total_claims
    )
# ==========================
# MANAGE STUDENTS
# ==========================

@app.route("/students")
def students():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            students.student_id,
            students.register_number,
            students.student_name,
            students.year_of_study,
            students.email,
            students.phone,
            departments.department_name
        FROM students
        LEFT JOIN departments
        ON students.department_id = departments.department_id
        ORDER BY students.student_id DESC
    """)

    students = cursor.fetchall()

    return render_template(
        "students.html",
        students=students
    )
    
@app.route("/add_student", methods=["GET", "POST"])
def add_student():

    if "admin_id" not in session:
        return redirect("/")

    if request.method == "POST":

        register_number = request.form["register_number"]
        student_name = request.form["student_name"]
        department_name = request.form["department"]
        year = request.form["year"]
        email = request.form["email"]
        phone = request.form["phone"]
        password = request.form["password"]

        cursor = mysql.connection.cursor()

        # Check department
        cursor.execute("""
            SELECT department_id
            FROM departments
            WHERE department_name=%s
        """, (department_name,))

        department = cursor.fetchone()

        if department is None:
            flash("Selected department not found!", "danger")
            return redirect("/add_student")

        department_id = department["department_id"]

        # Check duplicate register number
        cursor.execute("""
            SELECT student_id
            FROM students
            WHERE register_number=%s
        """, (register_number,))

        existing = cursor.fetchone()

        if existing:
            flash("Register Number already exists!", "danger")
            return redirect("/add_student")

        # Insert student
        cursor.execute("""
            INSERT INTO students
            (
                register_number,
                student_name,
                department_id,
                year_of_study,
                email,
                phone,
                password
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            register_number,
            student_name,
            department_id,
            year,
            email,
            phone,
            password
        ))

        mysql.connection.commit()

        flash("Student added successfully!", "success")

        return redirect("/students")

    return render_template("add_student.html")
@app.route("/edit_student/<int:student_id>", methods=["GET", "POST"])
def edit_student(student_id):

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    # ==========================
    # UPDATE STUDENT
    # ==========================
    if request.method == "POST":

        register_number = request.form["register_number"]
        student_name = request.form["student_name"]
        department = request.form["department"]
        year = request.form["year"]
        email = request.form["email"]
        phone = request.form["phone"]

        # Get Department ID
        cursor.execute("""
            SELECT department_id
            FROM departments
            WHERE department_name=%s
        """, (department,))

        dept = cursor.fetchone()

        if not dept:
            flash("Department not found!", "danger")
            return redirect("/edit_student/" + str(student_id))

        # Update Student
        cursor.execute("""
            UPDATE students
            SET register_number=%s,
                student_name=%s,
                department_id=%s,
                year_of_study=%s,
                email=%s,
                phone=%s
            WHERE student_id=%s
        """, (
            register_number,
            student_name,
            dept["department_id"],
            year,
            email,
            phone,
            student_id
        ))

        mysql.connection.commit()

        flash("Student Updated Successfully!", "success")

        # IMPORTANT
        return redirect("/students")

    # ==========================
    # GET STUDENT DETAILS
    # ==========================

    cursor.execute("""
        SELECT
            s.student_id,
            s.register_number,
            s.student_name,
            s.department_id,
            d.department_name,
            s.year_of_study,
            s.email,
            s.phone
        FROM students s
        LEFT JOIN departments d
            ON s.department_id = d.department_id
        WHERE s.student_id=%s
    """, (student_id,))

    student = cursor.fetchone()

    if not student:
        flash("Student not found!", "danger")
        return redirect("/students")

    # Get all departments
    cursor.execute("""
        SELECT department_id, department_name
        FROM departments
        ORDER BY department_name
    """)

    departments = cursor.fetchall()

    return render_template(
        "edit_student.html",
        student=student,
        departments=departments
    )
@app.route("/delete_student/<int:student_id>", methods=["GET", "POST"])
def delete_student(student_id):

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    try:
        # Delete related lost items
        cursor.execute("""
            DELETE FROM lost_items
            WHERE student_id = %s
        """, (student_id,))

        # Delete related found items
        cursor.execute("""
            DELETE FROM found_items
            WHERE student_id = %s
        """, (student_id,))

        # Delete student
        cursor.execute("""
            DELETE FROM students
            WHERE student_id = %s
        """, (student_id,))

        mysql.connection.commit()

        flash("Student Deleted Successfully!", "success")

    except Exception as e:

        mysql.connection.rollback()

        flash("Unable to delete student: " + str(e), "danger")

    return redirect("/students")

@app.route("/add_staff", methods=["GET", "POST"])
def add_staff():
    if request.method == "POST":
        # Staff details save code
        staff_code = request.form["staff_code"]
        staff_name = request.form["staff_name"]
        department = request.form["department"]
        email = request.form["email"]
        phone = request.form["phone"]
        password = request.form["password"]

        cursor = mysql.connection.cursor()

        cursor.execute(
    "SELECT department_id FROM departments WHERE department_name=%s",
    (department,)
)

        result = cursor.fetchone()
        print(result)

        if result is None:
           return "Department not found"

        department_id = result["department_id"]

        cursor.execute("""
          INSERT INTO staff
          (staff_code, staff_name, department_id, email, phone, password)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            staff_code,
            staff_name,
            department_id,
            email,
            phone,
            password
        ))

        mysql.connection.commit()

        flash("Staff added successfully!", "success")

        return redirect("/add_staff")

    return render_template("add_staff.html")

@app.route("/staff")
def staff():

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            s.staff_id,
            s.staff_code,
            s.staff_name,
            d.department_name,
            s.email,
            s.phone
        FROM staff s
        JOIN departments d
        ON s.department_id = d.department_id
    """)

    staff = cursor.fetchall()

    return render_template("staff.html", staff=staff)
@app.route("/edit_staff/<int:staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):

    cursor = mysql.connection.cursor()

    if request.method == "POST":

        staff_name = request.form["staff_name"]
        email = request.form["email"]
        phone = request.form["phone"]

        cursor.execute("""
            UPDATE staff
            SET staff_name=%s,
                email=%s,
                phone=%s
            WHERE staff_id=%s
        """, (staff_name, email, phone, staff_id))

        mysql.connection.commit()

        flash("Staff updated successfully!", "success")

        return redirect("/staff")

    cursor.execute("SELECT * FROM staff WHERE staff_id=%s", (staff_id,))
    staff = cursor.fetchone()

    return render_template("edit_staff.html", staff=staff)
@app.route("/delete_staff/<int:staff_id>")
def delete_staff(staff_id):

    cursor = mysql.connection.cursor()

    cursor.execute("DELETE FROM staff WHERE staff_id=%s", (staff_id,))

    mysql.connection.commit()

    flash("Staff deleted successfully!", "success")

    return redirect("/staff")
@app.route("/report_lost", methods=["GET", "POST"])
def report_lost():

    # Student or Admin must be logged in
    if "student_id" not in session and "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    # ==========================
    # POST - SAVE LOST ITEM
    # ==========================
    if request.method == "POST":

        # Student login
        if "student_id" in session:
            student_id = session["student_id"]

        # Admin login
        elif "admin_id" in session:
            student_id = request.form.get("student_id")

            if not student_id:
                flash("Please select a student!", "danger")
                return redirect("/report_lost")

        item_name = request.form.get("item_name")
        category = request.form.get("category")
        description = request.form.get("description")
        lost_location = request.form.get("lost_location")
        lost_date = request.form.get("lost_date")

        # ==========================
        # IMAGE UPLOAD
        # ==========================
        image = request.files.get("item_image")
        item_image = None

        if image and image.filename:
            upload_folder = os.path.join(
                app.static_folder,
                "uploads"
            )

            os.makedirs(upload_folder, exist_ok=True)

            filename = image.filename
            image_path = os.path.join(upload_folder, filename)

            image.save(image_path)

            item_image = "uploads/" + filename

        # ==========================
        # INSERT
        # ==========================
        cursor.execute("""
            INSERT INTO lost_items
            (
                student_id,
                item_name,
                category,
                description,
                lost_location,
                lost_date,
                item_image
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            student_id,
            item_name,
            category,
            description,
            lost_location,
            lost_date,
            item_image
        ))

        mysql.connection.commit()
        cursor.close()

        flash("Lost Item Reported Successfully!", "success")

        if "admin_id" in session:
            return redirect("/lost_items")

        return redirect("/my_lost_items")

    # ==========================
    # GET - ADMIN STUDENT LIST
    # ==========================
    students = []

    if "admin_id" in session:
        cursor.execute("""
            SELECT
                student_id,
                register_number,
                student_name
            FROM students
            ORDER BY student_name
        """)

        students = cursor.fetchall()

    cursor.close()

    return render_template(
        "report_lost.html",
        students=students
    )
@app.route("/lost_items")
def lost_items():

    if "student_id" not in session and "admin_id" not in session and "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("SELECT * FROM lost_items")
    items = cursor.fetchall()

    cursor.close()

    return render_template("lost_items.html", items=items)
@app.route("/my_lost_items")
def my_lost_items():

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT * FROM lost_items
        WHERE student_id=%s
        ORDER BY lost_id DESC
    """, (session["student_id"],))

    items = cursor.fetchall()

    return render_template("my_lost_items.html", items=items)

# ==========================
# LOGOUT
# ==========================

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")
@app.route("/report_found", methods=["GET", "POST"])
def report_found():

    if "student_id" not in session:
        return redirect("/")

    if request.method == "POST":

        student_id = session["student_id"]
        item_name = request.form["item_name"]
        category = request.form["category"]
        description = request.form["description"]
        found_location = request.form["found_location"]
        found_date = request.form["found_date"]
        image = request.files.get("item_image")

        item_image = None

        if image and image.filename:
            filename = image.filename

            image_path = os.path.join(
                app.static_folder,
                "uploads",
                filename
            )

            image.save(image_path)

            item_image = "uploads/" + filename
        
        cursor = mysql.connection.cursor()

        cursor.execute("""
            INSERT INTO found_items
            (student_id, item_name, category, description,
             found_location, found_date, item_image)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            student_id,
            item_name,
            category,
            description,
            found_location,
            found_date,
            item_image
        ))
        # Auto Match Lost Item

        cursor.execute("""
            UPDATE lost_items
            SET status='Found'
            WHERE student_id=%s
            AND item_name=%s
            AND category=%s
            AND status='Pending'
        """, (
    student_id,
    item_name,
    category
))
        mysql.connection.commit()

        flash("Found Item Reported Successfully!", "success")

        return redirect("/my_found_items")

    return render_template("report_found.html")

# ==========================
# MY FOUND ITEMS
# ==========================

@app.route("/my_found_items")
def my_found_items():

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT * FROM found_items
        WHERE student_id=%s
        ORDER BY found_id DESC
    """, (session["student_id"],))

    items = cursor.fetchall()

    return render_template("my_found_items.html", items=items)
@app.route("/my_profile")
def my_profile():

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            s.student_name,
            s.register_number,
            d.department_name,
            s.year_of_study,
            s.email,
            s.phone
        FROM students s
        JOIN departments d
        ON s.department_id = d.department_id
        WHERE s.student_id=%s
    """, (session["student_id"],))

    student = cursor.fetchone()

    return render_template("my_profile.html", student=student)
@app.route("/edit_profile", methods=["GET", "POST"])
def edit_profile():

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    if request.method == "POST":

        student_name = request.form["student_name"]
        email = request.form["email"]
        phone = request.form["phone"]

        cursor.execute("""
            UPDATE students
            SET student_name=%s,
                email=%s,
                phone=%s
            WHERE student_id=%s
        """, (
            student_name,
            email,
            phone,
            session["student_id"]
        ))

        mysql.connection.commit()

        session["student_name"] = student_name

        flash("Profile Updated Successfully!", "success")

        return redirect("/edit_profile")

    cursor.execute("""
        SELECT *
        FROM students
        WHERE student_id=%s
    """, (session["student_id"],))

    student = cursor.fetchone()

    return render_template("edit_profile.html", student=student)
@app.route("/change_password", methods=["GET", "POST"])
def change_password():

    if "student_id" not in session:
        return redirect("/")

    if request.method == "POST":

        current_password = request.form["current_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        cursor = mysql.connection.cursor()

        cursor.execute(
            "SELECT password FROM students WHERE student_id=%s",
            (session["student_id"],)
        )

        student = cursor.fetchone()

        if student["password"] != current_password:
            flash("Current Password is Incorrect!", "danger")
            return redirect("/change_password")

        if new_password != confirm_password:
            flash("New Password and Confirm Password do not match!", "danger")
            return redirect("/change_password")

        cursor.execute("""
            UPDATE students
            SET password=%s
            WHERE student_id=%s
        """, (
            new_password,
            session["student_id"]
        ))

        mysql.connection.commit()

        flash("Password Changed Successfully!", "success")

        return redirect("/change_password")

    return render_template("change_password.html")
@app.route("/claim_item/<int:lost_id>")
def claim_item(lost_id):

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    # Lost item details
    cursor.execute("""
        SELECT * FROM lost_items
        WHERE lost_id=%s
    """, (lost_id,))

    lost = cursor.fetchone()

    if not lost:
        flash("Lost item not found!", "danger")
        return redirect("/my_lost_items")

    # Matching found item
    cursor.execute("""
        SELECT * FROM found_items
        WHERE item_name=%s
        AND category=%s
        LIMIT 1
    """, (
        lost["item_name"],
        lost["category"]
    ))

    found = cursor.fetchone()

    print(lost)
    print(found)
    if not found:
        flash("No matching found item available.", "danger")
        return redirect("/my_lost_items")
    # Check if claim already exists
    cursor.execute("""
    SELECT * FROM claims
    WHERE lost_id=%s
    AND student_id=%s
""", (
    lost["lost_id"],
    session["student_id"]
    ))

    existing_claim = cursor.fetchone()

    if existing_claim:
        flash("You have already submitted a claim for this item.", "warning")
        return redirect("/my_lost_items")
    print("Before INSERT")
    # Insert claim
    cursor.execute("""
        INSERT INTO claims
        (lost_id, found_id, student_id)
        VALUES (%s,%s,%s)
    """, (
        lost["lost_id"],
        found["found_id"],
        session["student_id"]
    ))
    print("After INSERT")

    mysql.connection.commit()

    flash("Claim Request Sent Successfully!", "success")

    return redirect("/my_lost_items")
@app.route("/claim_requests")
def claim_requests():

    if "admin_id" not in session and "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            c.claim_id,
            s.student_name,
            s.register_number,
            l.item_name,
            l.category,
            c.claim_message,
            c.claim_status,
            c.request_date
        FROM claim_requests c
        INNER JOIN students s
            ON c.student_id = s.student_id
        INNER JOIN lost_items l
            ON c.lost_id = l.lost_id
        ORDER BY c.claim_id DESC
    """)

    claims = cursor.fetchall()

    return render_template(
        "claim_requests.html",
        claims=claims
    )
@app.route("/approve_claim/<int:claim_id>")
def approve_claim(claim_id):

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        UPDATE claim_requests
        SET claim_status = 'Approved'
        WHERE claim_id = %s
    """, (claim_id,))

    mysql.connection.commit()

    flash("Claim Approved Successfully!", "success")

    return redirect("/claim_requests")
@app.route("/reject_claim/<int:claim_id>")
def reject_claim(claim_id):

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        UPDATE claim_requests
        SET claim_status = 'Rejected'
        WHERE claim_id = %s
    """, (claim_id,))

    mysql.connection.commit()

    flash("Claim Rejected Successfully!", "danger")

    return redirect("/claim_requests")
@app.route("/my_claims")
def my_claims():

    if "student_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            c.claim_id,
            l.item_name,
            c.status
        FROM claims c
        INNER JOIN lost_items l
            ON c.lost_id = l.lost_id
        WHERE c.student_id=%s
        ORDER BY c.claim_id DESC
    """, (session["student_id"],))

    claims = cursor.fetchall()

    return render_template("my_claims.html", claims=claims)
@app.route("/verify_items")
def verify_items():

    if "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            found_id,
            item_name,
            category,
            found_location,
            found_date,
            status
        FROM found_items
        ORDER BY found_id DESC
    """)

    items = cursor.fetchall()

    return render_template("verify_items.html", items=items)
@app.route("/verify_found/<int:found_id>")
def verify_found(found_id):

    if "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        UPDATE found_items
        SET status='Verified'
        WHERE found_id=%s
    """, (found_id,))

    mysql.connection.commit()

    flash("Item Verified Successfully!", "success")

    return redirect("/verify_items")
@app.route("/department_items")
def department_items():

    if "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    # Staff department
    cursor.execute("""
        SELECT department_id
        FROM staff
        WHERE staff_id=%s
    """, (session["staff_id"],))

    staff = cursor.fetchone()

    department_id = staff["department_id"]

    # Department lost items
    cursor.execute("""
        SELECT
            l.lost_id,
            l.item_name,
            l.category,
            l.status,
            s.student_name
        FROM lost_items l
        JOIN students s
            ON l.student_id = s.student_id
        WHERE s.department_id=%s
    """, (department_id,))

    items = cursor.fetchall()

    return render_template("department_items.html", items=items)
@app.route("/reports")
def reports():

    if "staff_id" not in session and "admin_id" not in session:
        return redirect("/")

    return render_template("reports.html")

@app.route("/download_lost_report")
def download_lost_report():

    if "admin_id" not in session and "staff_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            l.lost_id,
            s.student_name,
            s.register_number,
            l.item_name,
            l.category,
            l.lost_location,
            l.lost_date,
            l.status
        FROM lost_items l
        JOIN students s
        ON l.student_id = s.student_id
        ORDER BY l.lost_id DESC
    """)

    items = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lost Items Report"

    ws.append([
        "Lost ID",
        "Student Name",
        "Register Number",
        "Item Name",
        "Category",
        "Location",
        "Lost Date",
        "Status"
    ])

    for item in items:
        ws.append([
            item["lost_id"],
            item["student_name"],
            item["register_number"],
            item["item_name"],
            item["category"],
            item["lost_location"],
            item["lost_date"],
            item["status"]
        ])

    # AUTOMATIC COLUMN WIDTH
    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="lost_items_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/download_students_report")
def download_students_report():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            s.student_id,
            s.register_number,
            s.student_name,
            d.department_name,
            s.year_of_study,
            s.email,
            s.phone,
            s.created_at
        FROM students s
        LEFT JOIN departments d
            ON s.department_id = d.department_id
        ORDER BY s.student_id DESC
    """)

    students = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append([
        "Student ID",
        "Register Number",
        "Student Name",
        "Department",
        "Year",
        "Email",
        "Phone",
        "Created Date"
    ])

    for student in students:
        ws.append([
            student["student_id"],
            student["register_number"],
            student["student_name"],
            student["department_name"],
            student["year_of_study"],
            student["email"],
            student["phone"],
            student["created_at"]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="students.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/download_staff_report")
def download_staff_report():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            s.staff_id,
            s.staff_code,
            s.staff_name,
            d.department_name,
            s.email,
            s.phone,
            s.created_at
        FROM staff s
        LEFT JOIN departments d
            ON s.department_id = d.department_id
        ORDER BY s.staff_id DESC
    """)

    staff_list = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"

    ws.append([
        "Staff ID",
        "Staff Code",
        "Staff Name",
        "Department",
        "Email",
        "Phone",
        "Created Date"
    ])

    for staff in staff_list:
        ws.append([
            staff["staff_id"],
            staff["staff_code"],
            staff["staff_name"],
            staff["department_name"],
            staff["email"],
            staff["phone"],
            staff["created_at"]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="staff.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/download_found_report")
def download_found_report():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            f.found_id,
            f.item_name,
            f.category,
            f.description,
            f.found_location,
            f.found_date,
            f.status,
            s.staff_name
        FROM found_items f
        LEFT JOIN staff s
            ON f.staff_id = s.staff_id
        ORDER BY f.found_id DESC
    """)

    items = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Found Items"

    ws.append([
        "Found ID",
        "Item Name",
        "Category",
        "Description",
        "Location",
        "Found Date",
        "Staff Name",
        "Status"
    ])

    for item in items:
        ws.append([
            item["found_id"],
            item["item_name"],
            item["category"],
            item["description"],
            item["found_location"],
            item["found_date"],
            item["staff_name"],
            item["status"]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="found_items.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/download_monthly_report")
def download_monthly_report():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            DATE_FORMAT(lost_date, '%%Y-%%m') AS month,
            COUNT(*) AS lost_items
        FROM lost_items
        WHERE lost_date IS NOT NULL
        GROUP BY DATE_FORMAT(lost_date, '%%Y-%%m')
        ORDER BY month DESC
    """)

    lost_data = cursor.fetchall()

    cursor.execute("""
        SELECT
            DATE_FORMAT(found_date, '%%Y-%%m') AS month,
            COUNT(*) AS found_items
        FROM found_items
        WHERE found_date IS NOT NULL
        GROUP BY DATE_FORMAT(found_date, '%%Y-%%m')
        ORDER BY month DESC
    """)

    found_data = cursor.fetchall()

    monthly = {}

    for row in lost_data:
        monthly[row["month"]] = {
            "lost": row["lost_items"],
            "found": 0
        }

    for row in found_data:
        if row["month"] not in monthly:
            monthly[row["month"]] = {
                "lost": 0,
                "found": 0
            }

        monthly[row["month"]]["found"] = row["found_items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    ws.append([
        "Month",
        "Lost Items",
        "Found Items",
        "Total"
    ])

    for month in sorted(monthly.keys(), reverse=True):
        lost_count = monthly[month]["lost"]
        found_count = monthly[month]["found"]

        ws.append([
            month,
            lost_count,
            found_count,
            lost_count + found_count
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="monthly_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/admins")
def admins():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT admin_id, username, full_name, created_at
        FROM admins
        ORDER BY admin_id DESC
    """)

    admins = cursor.fetchall()

    return render_template(
        "admins.html",
        admins=admins
    )
@app.route("/found_items")
def found_items():

    if "admin_id" not in session:
        return redirect("/")

    cursor = mysql.connection.cursor()

    cursor.execute("""
        SELECT
            f.found_id,
            f.item_name,
            f.category,
            f.description,
            f.found_location,
            f.found_date,
            f.status,
            s.student_name
        FROM found_items f
        LEFT JOIN students s
            ON f.student_id = s.student_id
        ORDER BY f.found_id DESC
    """)

    items = cursor.fetchall()

    return render_template(
        "found_items.html",
        items=items
    )


# ==========================
# RUN APPLICATION
# ==========================

if __name__ == "__main__":

    app.run(debug=True)